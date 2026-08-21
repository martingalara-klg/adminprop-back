"""Tests unitarios de `adminprop.modules.settlements.exports` -- RF-04
(agrupacion por propiedad) + RF-03 (Excel/PDF), sin DB (issue #30).

SDD: docs/sdd/features/spec_module_05_liquidaciones.md §RF-03/RF-04.
Implements: CA-05-07.
"""

from __future__ import annotations

import uuid
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook
from pypdf import PdfReader

from adminprop.modules.settlements.exports import (
    build_settlement_pdf,
    build_settlement_workbook,
    group_line_items_by_property,
)

_PROPERTY_A = uuid.uuid4()
_PROPERTY_B = uuid.uuid4()


@dataclass(frozen=True)
class _LineItem:
    line_type: str
    property_id: uuid.UUID | None
    original_amount: Decimal
    original_currency: str
    amount_ars: Decimal
    description: str | None = None


@dataclass(frozen=True)
class _Settlement:
    id: uuid.UUID
    period: date
    exchange_rate: Decimal | None
    total_collected: Decimal
    commission_total: Decimal
    charges_total: Decimal
    repairs_total: Decimal
    already_settled_total: Decimal
    net_amount: Decimal
    regenerated_count: int


def _settlement(**overrides) -> _Settlement:
    defaults = dict(
        id=uuid.uuid4(),
        period=date(2026, 6, 1),
        exchange_rate=None,
        total_collected=Decimal("180000.00"),
        commission_total=Decimal("18000.00"),
        charges_total=Decimal("5000.00"),
        repairs_total=Decimal("2000.00"),
        already_settled_total=Decimal("0.00"),
        net_amount=Decimal("155000.00"),
        regenerated_count=0,
    )
    defaults.update(overrides)
    return _Settlement(**defaults)


class TestGroupLineItemsByProperty:
    """RF-04: "cada propiedad con sus cobros, cargos y reparaciones, con
    subtotal; el consolidado del propietario al final"."""

    def test_groups_items_by_property_and_computes_subtotal(self):
        items = [
            _LineItem("rent_collected", _PROPERTY_A, Decimal("100000"), "ARS", Decimal("100000.00")),
            _LineItem("tax_charge", _PROPERTY_A, Decimal("5000"), "ARS", Decimal("5000.00")),
            _LineItem("rent_collected", _PROPERTY_B, Decimal("80000"), "ARS", Decimal("80000.00")),
            _LineItem("repair", _PROPERTY_B, Decimal("2000"), "ARS", Decimal("2000.00")),
            _LineItem("commission", None, Decimal("18000"), "ARS", Decimal("18000.00")),
        ]
        labels = {_PROPERTY_A: "Av. Test 123", _PROPERTY_B: "Belgrano 456"}

        groups, general_items = group_line_items_by_property(items, labels)

        assert len(groups) == 2
        assert len(general_items) == 1
        assert general_items[0].line_type == "commission"

        group_a = next(g for g in groups if g.property_id == _PROPERTY_A)
        assert group_a.property_label == "Av. Test 123"
        assert group_a.subtotal_ars == Decimal("95000.00")  # 100000 - 5000

        group_b = next(g for g in groups if g.property_id == _PROPERTY_B)
        assert group_b.subtotal_ars == Decimal("78000.00")  # 80000 - 2000

    def test_already_settled_is_listed_but_not_in_subtotal(self):
        """RN-L01/RN-P07: "ya rendido" es informativo -- no afecta el
        subtotal de la propiedad."""
        items = [
            _LineItem(
                "already_settled", _PROPERTY_A, Decimal("50000"), "ARS", Decimal("50000.00")
            ),
        ]
        groups, _general = group_line_items_by_property(items, {_PROPERTY_A: "Prop A"})

        assert len(groups) == 1
        assert groups[0].subtotal_ars == Decimal("0.00")
        assert len(groups[0].line_items) == 1

    def test_unknown_property_id_falls_back_to_raw_uuid_label(self):
        items = [
            _LineItem("rent_collected", _PROPERTY_A, Decimal("1000"), "ARS", Decimal("1000.00")),
        ]
        groups, _general = group_line_items_by_property(items, {})
        assert groups[0].property_label == str(_PROPERTY_A)

    def test_groups_sorted_by_property_label(self):
        items = [
            _LineItem("rent_collected", _PROPERTY_B, Decimal("1"), "ARS", Decimal("1.00")),
            _LineItem("rent_collected", _PROPERTY_A, Decimal("1"), "ARS", Decimal("1.00")),
        ]
        labels = {_PROPERTY_A: "AAA", _PROPERTY_B: "ZZZ"}
        groups, _general = group_line_items_by_property(items, labels)
        assert [g.property_label for g in groups] == ["AAA", "ZZZ"]


class TestBuildSettlementWorkbook:
    """RF-03/CA-05-07: Excel releido con openpyxl (nunca byte a byte)."""

    def test_workbook_contains_property_sections_and_consolidated_summary(self):
        items = [
            _LineItem("rent_collected", _PROPERTY_A, Decimal("100000"), "ARS", Decimal("100000.00")),
            _LineItem("commission", None, Decimal("10000"), "ARS", Decimal("10000.00")),
        ]
        groups, general_items = group_line_items_by_property(items, {_PROPERTY_A: "Av. Test 123"})
        settlement = _settlement()

        content = build_settlement_workbook(
            settlement=settlement,
            landlord_name="Juan Perez",
            property_groups=groups,
            general_items=general_items,
        )

        workbook = load_workbook(BytesIO(content))
        sheet = workbook.active
        cell_values = [cell.value for row in sheet.iter_rows() for cell in row if cell.value]
        joined = " ".join(str(v) for v in cell_values)
        assert "Juan Perez" in joined
        assert "Av. Test 123" in joined
        assert "Subtotal propiedad" in joined
        assert "Consolidado del propietario" in joined
        assert "Neto a rendir" in joined


class TestBuildSettlementPdf:
    """RF-03/RF-04/CA-05-07: PDF releido con pypdf (extraccion de texto)."""

    def test_pdf_contains_property_sections_and_consolidated_summary(self):
        items = [
            _LineItem("rent_collected", _PROPERTY_A, Decimal("100000"), "ARS", Decimal("100000.00")),
            _LineItem("commission", None, Decimal("10000"), "ARS", Decimal("10000.00")),
        ]
        groups, general_items = group_line_items_by_property(items, {_PROPERTY_A: "Av. Test 123"})
        settlement = _settlement()

        pdf_bytes = build_settlement_pdf(
            settlement=settlement,
            landlord_name="Juan Perez",
            property_groups=groups,
            general_items=general_items,
            billing_header={"name": "Administradora Test", "cuit": "30-12345678-9"},
        )

        assert pdf_bytes.startswith(b"%PDF-")
        reader = PdfReader(BytesIO(pdf_bytes))
        text = "\n".join(page.extract_text() for page in reader.pages)
        assert "Administradora Test" in text
        assert "Av. Test 123" in text
        assert "Consolidado del propietario" in text
