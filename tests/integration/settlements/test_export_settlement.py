"""tests/integration/settlements/test_export_settlement.py -- issue #30.

SDD: docs/sdd/features/spec_module_05_liquidaciones.md §RF-03/RF-04 +
core/sdd_03_api_contracts.md §11 "GET /settlements/:id/export?format=".
Implements: CA-05-07.
"""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO
from unittest.mock import MagicMock

import pytest

from adminprop.db.session import get_session_factory, set_tenant_context
from adminprop.modules.settlements.repository import SettlementRepository
from adminprop.workers.documents_worker import _generate_settlement_async

pytestmark = pytest.mark.asyncio


async def _seed_generated_settlement(seed, *, period="2026-06"):
    """A diferencia de otros tests HTTP, aca se corre el calculo REAL
    (`_generate_settlement_async`, no mockeado) porque los exports son un
    efecto secundario del worker -- sin esto no habria adjuntos que
    descargar."""
    org = await seed.create_organization_with_system_roles()
    owner = await seed.add_member(
        organization_id=org["organization_id"], role_id=org["roles"]["owner"], role_name="owner"
    )
    landlord_id = await seed.create_landlord_row(organization_id=org["organization_id"])
    renter_id = await seed.create_renter_row(organization_id=org["organization_id"])
    property_id = await seed.create_property_row(
        organization_id=org["organization_id"], landlord_id=landlord_id, address="Av. Colon 100"
    )
    contract_id = await seed.create_contract_row(
        organization_id=org["organization_id"], property_id=property_id, renter_id=renter_id
    )
    rent_period_id = await seed.create_rent_period_row(
        organization_id=org["organization_id"], contract_id=contract_id, period=f"{period}-01"
    )
    await seed.create_payment_row(
        organization_id=org["organization_id"],
        rent_period_id=rent_period_id,
        created_by=owner["id"],
        amount="80000.00",
    )

    session_factory = get_session_factory()
    async with session_factory() as session, session.begin():
        await set_tenant_context(session, org["organization_id"])
        repo = SettlementRepository(session)
        settlement = await repo.create_placeholder(
            organization_id=org["organization_id"],
            landlord_id=landlord_id,
            period=date.fromisoformat(f"{period}-01"),
            exchange_rate=None,
            commission_pct_used=Decimal("10.00"),
            generated_by=owner["id"],
        )
        settlement_id = settlement.id

    await _generate_settlement_async(settlement_id, org["organization_id"], "req-export-test")
    return org, owner, settlement_id


class TestExportSettlement:
    """CA-05-07: "el export Excel y el PDF ... quedan descargables desde
    el detalle"."""

    async def test_ca_05_07_export_xlsx_downloads_valid_workbook(self, client, seed):
        _org, owner, settlement_id = await _seed_generated_settlement(seed)

        response = await client.get(
            f"/v1/settlements/{settlement_id}/export?format=xlsx", headers=owner["headers"]
        )

        assert response.status_code == 200
        assert (
            response.headers["content-type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "attachment" in response.headers["content-disposition"]

        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(response.content))
        assert workbook.active.max_row > 0

    async def test_export_pdf_downloads_valid_pdf(self, client, seed):
        _org, owner, settlement_id = await _seed_generated_settlement(seed)

        response = await client.get(
            f"/v1/settlements/{settlement_id}/export?format=pdf", headers=owner["headers"]
        )

        assert response.status_code == 200
        assert response.headers["content-type"] == "application/pdf"
        assert response.content.startswith(b"%PDF-")

    async def test_export_appears_listed_in_settlement_detail(self, client, seed):
        _org, owner, settlement_id = await _seed_generated_settlement(seed)

        response = await client.get(f"/v1/settlements/{settlement_id}", headers=owner["headers"])

        assert response.status_code == 200
        attachments = response.json()["data"]["attachments"]
        formats = {a["format"] for a in attachments}
        assert formats == {"xlsx", "pdf"}

    async def test_export_invalid_format_returns_400_validation_error(self, client, seed):
        _org, owner, settlement_id = await _seed_generated_settlement(seed)

        response = await client.get(
            f"/v1/settlements/{settlement_id}/export?format=docx", headers=owner["headers"]
        )

        assert response.status_code == 400
        assert response.json()["error"]["code"] == "VALIDATION_ERROR"

    async def test_export_without_generated_documents_returns_404(self, client, seed, monkeypatch):
        """Liquidacion existe pero el job de calculo (y sus exports)
        todavia no corrio -- `404 NOT_FOUND`."""
        monkeypatch.setattr(
            "adminprop.workers.documents_worker.generate_settlement.apply_async", MagicMock()
        )
        org = await seed.create_organization_with_system_roles()
        owner = await seed.add_member(
            organization_id=org["organization_id"],
            role_id=org["roles"]["owner"],
            role_name="owner",
        )
        landlord_id = await seed.create_landlord_row(organization_id=org["organization_id"])
        renter_id = await seed.create_renter_row(organization_id=org["organization_id"])
        property_id = await seed.create_property_row(
            organization_id=org["organization_id"], landlord_id=landlord_id
        )
        contract_id = await seed.create_contract_row(
            organization_id=org["organization_id"], property_id=property_id, renter_id=renter_id
        )
        await seed.create_rent_period_row(
            organization_id=org["organization_id"], contract_id=contract_id, period="2026-06-01"
        )

        generate_response = await client.post(
            "/v1/settlements/generate",
            json={"landlord_id": str(landlord_id), "period": "2026-06"},
            headers=owner["headers"],
        )
        assert generate_response.status_code == 202
        settlement_id = generate_response.json()["data"]["settlement_id"]

        response = await client.get(
            f"/v1/settlements/{settlement_id}/export?format=pdf", headers=owner["headers"]
        )

        assert response.status_code == 404
        assert response.json()["error"]["code"] == "NOT_FOUND"

    async def test_export_of_other_tenant_settlement_returns_404(self, client, seed):
        org_a = await seed.create_organization_with_system_roles()
        owner_a = await seed.add_member(
            organization_id=org_a["organization_id"],
            role_id=org_a["roles"]["owner"],
            role_name="owner",
        )
        _org_b, _owner_b, settlement_b_id = await _seed_generated_settlement(seed, period="2026-07")

        response = await client.get(
            f"/v1/settlements/{settlement_b_id}/export?format=pdf", headers=owner_a["headers"]
        )

        assert response.status_code == 404
        assert response.json()["error"]["code"] == "NOT_FOUND"
