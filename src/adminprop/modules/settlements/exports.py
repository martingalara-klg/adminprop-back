"""Agrupacion por propiedad (RF-04) + generacion de los exports Excel/PDF
de una liquidacion (RF-03) -- issue #30.

SDD: docs/sdd/features/spec_module_05_liquidaciones.md §RF-03 ("Excel
(openpyxl) + PDF (WeasyPrint) generados por `documents_worker`") + §RF-04
("el detalle y los exports agrupan por propiedad: cada propiedad con sus
cobros, cargos y reparaciones, con subtotal; el consolidado del
propietario al final").

`group_line_items_by_property` es una funcion PURA (sin I/O) -- mismo
criterio que `service.calculate_settlement`: reutilizable tanto por
`GET /settlements/:id?scope=per_property` (schemas puros) como por
`build_settlement_workbook`/`build_settlement_pdf` (bytes), y
unit-testeable sin Postgres real. Acepta cualquier objeto con los atributos
de una linea (duck typing): sirve igual para instancias ORM
(`SettlementLineItem`, dentro del worker) que para los schemas Pydantic ya
serializados (`SettlementLineItemDetail`, en el endpoint de detalle).
"""

from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal
from io import BytesIO
from typing import Any, Protocol
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font

from adminprop.shared.pdf import DocumentSection, document_html_multi_section, render_pdf_from_html

# RF-02: tipos que restan del neto a rendir de una propiedad (junto con
# `rent_collected`, que suma) -- `commission` no es por propiedad (RN-L02,
# se calcula sobre TODAS las propiedades del dueño) y `already_settled` es
# puramente informativo (RN-L01/RN-P07, no participa del neto).
_SUBTRACTING_LINE_TYPES = frozenset({"tax_charge", "repair"})
_ADDING_LINE_TYPES = frozenset({"rent_collected"})

# Etiquetas legibles para el export -- CA-05-07 (Excel/PDF descargables).
_LINE_TYPE_LABELS: dict[str, str] = {
    "rent_collected": "Alquiler cobrado",
    "already_settled": "Ya rendido (cobro directo del propietario)",
    "tax_charge": "Cargo del mes",
    "repair": "Reparacion (agency)",
    "commission": "Comision de administracion",
}


class LineItemLike(Protocol):
    line_type: str
    property_id: UUID | None
    original_amount: Decimal
    original_currency: str
    amount_ars: Decimal
    description: str | None


@dataclass(frozen=True)
class PropertyGroup:
    """RF-04: una propiedad con sus lineas y el subtotal (RN-L01: suma
    `rent_collected`, resta `tax_charge`/`repair`; `already_settled` queda
    listado pero no afecta el subtotal, igual que en el neto general)."""

    property_id: UUID
    property_label: str
    line_items: list[Any]
    subtotal_ars: Decimal


def group_line_items_by_property(
    line_items: list[Any], property_labels: dict[UUID, str]
) -> tuple[list[PropertyGroup], list[Any]]:
    """Separa las lineas con `property_id` (agrupadas y ordenadas por
    direccion) de las que no lo tienen (`commission`, siempre a nivel del
    propietario -- van al consolidado, no a ninguna propiedad). Devuelve
    `(grupos_por_propiedad, lineas_generales)`."""
    by_property: dict[UUID, list[Any]] = {}
    general_items: list[Any] = []

    for item in line_items:
        if item.property_id is None:
            general_items.append(item)
        else:
            by_property.setdefault(item.property_id, []).append(item)

    groups: list[PropertyGroup] = []
    for property_id, items in by_property.items():
        subtotal = Decimal("0.00")
        for item in items:
            if item.line_type in _ADDING_LINE_TYPES:
                subtotal += item.amount_ars
            elif item.line_type in _SUBTRACTING_LINE_TYPES:
                subtotal -= item.amount_ars
        groups.append(
            PropertyGroup(
                property_id=property_id,
                property_label=property_labels.get(property_id, str(property_id)),
                line_items=items,
                subtotal_ars=subtotal,
            )
        )

    groups.sort(key=lambda g: g.property_label)
    return groups, general_items


def _format_ars(amount: Decimal) -> str:
    return f"$ {amount:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")


def _line_item_row(item: Any) -> tuple[str, str]:
    label = _LINE_TYPE_LABELS.get(item.line_type, item.line_type)
    if item.description:
        label = f"{label} -- {item.description}"
    if item.original_currency != "ARS":
        value = f"{_format_ars(item.amount_ars)} (orig. {item.original_amount} {item.original_currency})"
    else:
        value = _format_ars(item.amount_ars)
    return label, value


def _settlement_summary_rows(settlement: Any) -> list[tuple[str, str]]:
    return [
        ("Periodo", settlement.period.strftime("%Y-%m")),
        ("Cobrado (destino administracion)", _format_ars(settlement.total_collected)),
        ("Comision de administracion", _format_ars(settlement.commission_total)),
        ("Cargos del mes", _format_ars(settlement.charges_total)),
        ("Reparaciones (agency)", _format_ars(settlement.repairs_total)),
        ("Ya rendido (informativo)", _format_ars(settlement.already_settled_total)),
        ("Tipo de cambio", str(settlement.exchange_rate) if settlement.exchange_rate else "N/A"),
        ("Regeneraciones", str(settlement.regenerated_count)),
        ("Neto a rendir", _format_ars(settlement.net_amount)),
    ]


def build_settlement_pdf(
    *,
    settlement: Any,
    landlord_name: str,
    property_groups: list[PropertyGroup],
    general_items: list[Any],
    billing_header: dict,
) -> bytes:
    """RF-03/RF-04: PDF agrupado por propiedad, con el consolidado del
    propietario al final (mismo orden que pide RF-04). Usa el generador
    compartido `shared/pdf/template.py.document_html_multi_section`
    (extension del generador de `document_html`, ver docstring de ese
    modulo)."""
    sections: list[DocumentSection] = []
    for group in property_groups:
        sections.append(
            DocumentSection(
                heading=group.property_label,
                rows=[_line_item_row(item) for item in group.line_items],
                subtotal=("Subtotal propiedad", _format_ars(group.subtotal_ars)),
            )
        )

    consolidated_rows = [_line_item_row(item) for item in general_items]
    consolidated_rows.extend(_settlement_summary_rows(settlement))
    sections.append(
        DocumentSection(
            heading="Consolidado del propietario",
            rows=consolidated_rows,
            subtotal=("Neto a rendir", _format_ars(settlement.net_amount)),
        )
    )

    html = document_html_multi_section(
        title=f"Liquidacion -- {landlord_name} -- {settlement.period.strftime('%Y-%m')}",
        billing_header=billing_header,
        sections=sections,
        footer=f"Liquidacion {settlement.id} -- generada por AdminProp.",
    )
    return render_pdf_from_html(html)


def build_settlement_workbook(
    *,
    settlement: Any,
    landlord_name: str,
    property_groups: list[PropertyGroup],
    general_items: list[Any],
) -> bytes:
    """RF-03/RF-04: Excel (openpyxl) agrupado por propiedad con
    subtotales, consolidado al final -- releido/verificado por los tests
    con `openpyxl.load_workbook` (nunca comparacion byte a byte, mismo
    criterio que `pypdf` para el PDF)."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Liquidacion"

    bold = Font(bold=True)
    row_index = 1

    sheet.cell(row=row_index, column=1, value=f"Liquidacion -- {landlord_name}").font = bold
    row_index += 1
    sheet.cell(row=row_index, column=1, value=f"Periodo: {settlement.period.strftime('%Y-%m')}")
    row_index += 2

    for group in property_groups:
        sheet.cell(row=row_index, column=1, value=group.property_label).font = bold
        row_index += 1
        sheet.cell(row=row_index, column=1, value="Concepto").font = bold
        sheet.cell(row=row_index, column=2, value="Moneda orig.").font = bold
        sheet.cell(row=row_index, column=3, value="Monto orig.").font = bold
        sheet.cell(row=row_index, column=4, value="Monto ARS").font = bold
        row_index += 1
        for item in group.line_items:
            label = _LINE_TYPE_LABELS.get(item.line_type, item.line_type)
            if item.description:
                label = f"{label} -- {item.description}"
            sheet.cell(row=row_index, column=1, value=label)
            sheet.cell(row=row_index, column=2, value=item.original_currency)
            sheet.cell(row=row_index, column=3, value=float(item.original_amount))
            sheet.cell(row=row_index, column=4, value=float(item.amount_ars))
            row_index += 1
        sheet.cell(row=row_index, column=1, value="Subtotal propiedad").font = bold
        sheet.cell(row=row_index, column=4, value=float(group.subtotal_ars)).font = bold
        row_index += 2

    sheet.cell(row=row_index, column=1, value="Consolidado del propietario").font = bold
    row_index += 1
    for item in general_items:
        label = _LINE_TYPE_LABELS.get(item.line_type, item.line_type)
        sheet.cell(row=row_index, column=1, value=label)
        sheet.cell(row=row_index, column=4, value=float(item.amount_ars))
        row_index += 1

    for label, value in _settlement_summary_rows(settlement):
        sheet.cell(row=row_index, column=1, value=label)
        sheet.cell(row=row_index, column=4, value=value)
        row_index += 1

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
